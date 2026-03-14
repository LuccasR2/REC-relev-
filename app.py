import streamlit as st
import pandas as pd
import numpy as np
from datetime import datetime, date
import io
from copy import deepcopy

st.set_page_config(
    page_title="Relevé de Compte Copropriété",
    page_icon="🏢",
    layout="wide"
)

st.markdown("""
<style>
    .main-title { font-size: 1.6rem; font-weight: 700; color: #1a3c6e; margin-bottom: 0.2rem; }
    .sub-title { font-size: 0.95rem; color: #555; margin-bottom: 1.5rem; }
    .section-header { font-size: 1.1rem; font-weight: 600; color: #1a3c6e; border-bottom: 2px solid #1a3c6e; padding-bottom: 4px; margin-bottom: 1rem; }
    .solde-positif { color: #c0392b; font-weight: bold; }
    .solde-negatif { color: #27ae60; font-weight: bold; }
    .credit-row { background-color: #eafaf1; }
    .debit-row { background-color: #fef9e7; }
    .deleted-row { opacity: 0.4; text-decoration: line-through; }
    .stDataFrame { font-size: 0.85rem; }
    div[data-testid="metric-container"] { background: #f0f4fb; border-radius: 8px; padding: 10px; }
</style>
""", unsafe_allow_html=True)


# ─── helpers ──────────────────────────────────────────────────────────────────

def parse_date(val):
    if pd.isna(val):
        return None
    if isinstance(val, (datetime, date)):
        return pd.Timestamp(val)
    try:
        return pd.Timestamp(str(val), dayfirst=True)
    except Exception:
        return None


def load_excel(uploaded_file):
    xl = pd.ExcelFile(uploaded_file)
    dfs = {}
    for sheet in xl.sheet_names:
        dfs[sheet] = pd.read_excel(uploaded_file, sheet_name=sheet, header=None)
    return dfs


def parse_releve(df_raw):
    """Parse the 'Relevé imputé' sheet (skip first 3 header rows)."""
    header_row = None
    for i, row in df_raw.iterrows():
        vals = [str(v).strip() for v in row if not pd.isna(v)]
        if "Date" in vals and "Libellé" in vals:
            header_row = i
            break
    if header_row is None:
        header_row = 3
    df = df_raw.iloc[header_row + 1:].copy()
    df.columns = ["Date", "Libellé", "Débit (€)", "Crédit (€)", "Imputé sur", "Surplus (€)", "Solde (€)"]
    df = df[df["Date"].notna() | df["Libellé"].notna()].copy()
    df["Date"] = df["Date"].apply(parse_date)
    df["Débit (€)"] = pd.to_numeric(df["Débit (€)"], errors="coerce")
    df["Crédit (€)"] = pd.to_numeric(df["Crédit (€)"], errors="coerce")
    df["Solde (€)"] = pd.to_numeric(df["Solde (€)"], errors="coerce")
    df = df.reset_index(drop=True)
    return df


def parse_dettes(df_raw):
    header_row = None
    for i, row in df_raw.iterrows():
        vals = [str(v).strip() for v in row if not pd.isna(v)]
        if "Date" in vals and "Libellé" in vals:
            header_row = i
            break
    if header_row is None:
        header_row = 1
    df = df_raw.iloc[header_row + 1:].copy()
    df.columns = ["Date", "Libellé", "Montant initial (€)", "Solde restant (€)"]
    df = df[df["Date"].notna() | df["Libellé"].notna()].copy()
    df["Date"] = df["Date"].apply(parse_date)
    df["Montant initial (€)"] = pd.to_numeric(df["Montant initial (€)"], errors="coerce")
    df["Solde restant (€)"] = pd.to_numeric(df["Solde restant (€)"], errors="coerce")
    df = df.reset_index(drop=True)
    return df


def imputer_reglements(df_in, active_mask):
    """
    Re-calcule l'imputation des règlements sur la dette la plus ancienne
    en tenant compte des lignes actives/supprimées.
    """
    df = df_in.copy()
    # On ne recalcule que les lignes actives
    active_idx = [i for i, a in enumerate(active_mask) if a]

    # Construire la file de dettes (débits) triée chronologiquement
    debts = []
    for i in active_idx:
        row = df.iloc[i]
        if pd.notna(row["Débit (€)"]) and pd.isna(row["Crédit (€)"]):
            debts.append({
                "idx": i,
                "date": row["Date"],
                "libelle": row["Libellé"],
                "montant": row["Débit (€)"],
                "reste": row["Débit (€)"],
            })
    # Trier par date
    debts.sort(key=lambda x: x["date"] if x["date"] is not None else pd.Timestamp("2099-01-01"))

    # Réinitialiser la colonne "Imputé sur" pour les règlements
    imputation_map = {}

    # Appliquer les règlements (crédits) dans l'ordre chronologique
    credit_rows = [(i, df.iloc[i]) for i in active_idx
                   if pd.notna(df.iloc[i]["Crédit (€)"]) and pd.isna(df.iloc[i]["Débit (€)"])]
    credit_rows.sort(key=lambda x: x[1]["Date"] if x[1]["Date"] is not None else pd.Timestamp("2099-01-01"))

    for cr_i, cr_row in credit_rows:
        montant_credit = cr_row["Crédit (€)"]
        reste_credit = montant_credit
        imputs = []
        for debt in debts:
            if reste_credit <= 0:
                break
            if debt["reste"] <= 0:
                continue
            applique = min(reste_credit, debt["reste"])
            debt["reste"] -= applique
            reste_credit -= applique
            imputs.append(f"{debt['libelle']} ({applique:.2f}€)")
        imputation_map[cr_i] = "; ".join(imputs) if imputs else ""

    # Recalculer les soldes
    solde = 0.0
    soldes = {}
    for i in active_idx:
        row = df.iloc[i]
        if pd.notna(row["Débit (€)"]) and pd.isna(row["Crédit (€)"]):
            solde += row["Débit (€)"]
        elif pd.notna(row["Crédit (€)"]) and pd.isna(row["Débit (€)"]):
            solde -= row["Crédit (€)"]
        soldes[i] = solde

    return imputation_map, soldes, debts


def get_copro_name(df_raw):
    try:
        first_row = " ".join(str(v) for v in df_raw.iloc[0] if not pd.isna(v))
        if "RELEVÉ DE COMPTE" in first_row and "—" in first_row:
            return first_row.split("—")[-1].strip()
    except Exception:
        pass
    return "Copropriétaire"


def format_montant(val):
    if pd.isna(val) or val is None:
        return ""
    return f"{val:,.2f} €".replace(",", " ")


# ─── session state ─────────────────────────────────────────────────────────────

def init_state():
    defaults = {
        "releve_df": None,
        "dettes_df": None,
        "active_mask": None,
        "copro_name": "",
        "file_loaded": False,
        "new_rows": [],
    }
    for k, v in defaults.items():
        if k not in st.session_state:
            st.session_state[k] = v


init_state()

# ─── sidebar ───────────────────────────────────────────────────────────────────

with st.sidebar:
    st.image("https://img.icons8.com/color/96/building.png", width=60)
    st.markdown("### Relevé Copropriété")
    st.markdown("---")

    uploaded = st.file_uploader(
        "📂 Importer un relevé Excel",
        type=["xlsx"],
        help="Fichier .xlsx avec les feuilles 'Relevé imputé' et 'Dettes non soldées'"
    )

    if uploaded:
        try:
            dfs = load_excel(uploaded)
            sheets = list(dfs.keys())

            sheet_releve = sheets[0] if len(sheets) > 0 else None
            sheet_dettes = sheets[1] if len(sheets) > 1 else None

            if sheet_releve:
                df_rel = parse_releve(dfs[sheet_releve])
                st.session_state["releve_df"] = df_rel
                st.session_state["active_mask"] = [True] * len(df_rel)
                st.session_state["copro_name"] = get_copro_name(dfs[sheet_releve])

            if sheet_dettes:
                st.session_state["dettes_df"] = parse_dettes(dfs[sheet_dettes])

            st.session_state["file_loaded"] = True
            st.session_state["new_rows"] = []
            st.success("✅ Fichier chargé avec succès")
        except Exception as e:
            st.error(f"Erreur de lecture : {e}")

    st.markdown("---")
    if st.session_state["file_loaded"]:
        st.markdown(f"**Copropriétaire :** {st.session_state['copro_name']}")
        df_r = st.session_state["releve_df"]
        n_active = sum(st.session_state["active_mask"])
        st.markdown(f"**Lignes actives :** {n_active} / {len(df_r)}")

    st.markdown("---")
    st.markdown("#### ℹ️ Mode d'emploi")
    st.markdown("""
- Importez un relevé Excel
- Cochez/décochez les lignes
- Ajoutez des lignes manuellement
- Exportez le relevé recalculé
""")


# ─── main ──────────────────────────────────────────────────────────────────────

st.markdown('<div class="main-title">🏢 Relevé de Compte — Copropriété</div>', unsafe_allow_html=True)
st.markdown('<div class="sub-title">Imputation des règlements sur la dette la plus ancienne</div>', unsafe_allow_html=True)

if not st.session_state["file_loaded"]:
    st.info("👈 Commencez par importer un fichier Excel dans le panneau latéral.")
    st.markdown("""
    ### Format attendu
    Le fichier Excel doit contenir deux feuilles :
    - **Relevé imputé** : Date, Libellé, Débit (€), Crédit (€), Imputé sur, Surplus (€), Solde (€)
    - **Dettes non soldées** : Date, Libellé, Montant initial (€), Solde restant (€)
    """)
    st.stop()

df_rel = st.session_state["releve_df"]
active_mask = st.session_state["active_mask"]
copro_name = st.session_state["copro_name"]

# ── Tabs ───────────────────────────────────────────────────────────────────────
tab1, tab2, tab3 = st.tabs(["📋 Relevé de compte", "📊 Dettes non soldées", "➕ Ajouter des lignes"])

# ═══════════════════════════════════════════════════════════════════
# TAB 1 — Relevé de compte
# ═══════════════════════════════════════════════════════════════════
with tab1:
    col_head1, col_head2 = st.columns([3, 1])
    with col_head1:
        st.markdown(f"**Copropriétaire :** {copro_name}")
        st.markdown(f"*Règlement par extinction de la dette la plus ancienne — généré le {datetime.now().strftime('%d/%m/%Y')}*")
    with col_head2:
        n_active = sum(active_mask)
        st.metric("Lignes actives", f"{n_active} / {len(df_rel)}")

    # ── Filtres ─────────────────────────────────────────────────────
    with st.expander("🔍 Filtres", expanded=False):
        col_f1, col_f2 = st.columns(2)
        with col_f1:
            search = st.text_input("Rechercher dans le libellé", "")
        with col_f2:
            type_filter = st.selectbox("Type d'opération", ["Tous", "Débits uniquement", "Crédits uniquement"])

    # ── Recalcul ────────────────────────────────────────────────────
    imputation_map, soldes_map, debts_remaining = imputer_reglements(df_rel, active_mask)

    # ── Affichage ligne par ligne ────────────────────────────────────
    st.markdown('<div class="section-header">Lignes du relevé</div>', unsafe_allow_html=True)

    col_check, col_date, col_lib, col_debit, col_credit, col_impute, col_solde = st.columns(
        [0.4, 1.0, 2.8, 0.9, 0.9, 2.5, 0.9]
    )
    col_check.markdown("**✓**")
    col_date.markdown("**Date**")
    col_lib.markdown("**Libellé**")
    col_debit.markdown("**Débit (€)**")
    col_credit.markdown("**Crédit (€)**")
    col_impute.markdown("**Imputé sur**")
    col_solde.markdown("**Solde (€)**")

    st.markdown("---")

    for i, row in df_rel.iterrows():
        # Apply filters
        libelle_str = str(row["Libellé"]) if not pd.isna(row["Libellé"]) else ""
        if search and search.lower() not in libelle_str.lower():
            continue
        is_debit = pd.notna(row["Débit (€)"]) and pd.isna(row["Crédit (€)"])
        is_credit = pd.notna(row["Crédit (€)"]) and pd.isna(row["Débit (€)"])
        if type_filter == "Débits uniquement" and not is_debit:
            continue
        if type_filter == "Crédits uniquement" and not is_credit:
            continue

        current_active = active_mask[i]
        solde_val = soldes_map.get(i, None)
        impute_str = imputation_map.get(i, "")
        if pd.isna(row.get("Imputé sur")) or str(row.get("Imputé sur", "")) in ("nan", ""):
            impute_orig = impute_str
        else:
            # Keep original for non-credit rows
            impute_orig = "" if is_credit else str(row.get("Imputé sur", ""))

        col_check, col_date, col_lib, col_debit, col_credit, col_impute, col_solde = st.columns(
            [0.4, 1.0, 2.8, 0.9, 0.9, 2.5, 0.9]
        )
        with col_check:
            new_val = st.checkbox(
                "", value=current_active, key=f"active_{i}",
                label_visibility="collapsed"
            )
            if new_val != active_mask[i]:
                st.session_state["active_mask"][i] = new_val
                st.rerun()

        opacity = "1" if current_active else "0.4"
        td = row["Débit (€)"] if pd.notna(row["Débit (€)"]) else None
        tc = row["Crédit (€)"] if pd.notna(row["Crédit (€)"]) else None

        with col_date:
            date_str = row["Date"].strftime("%d/%m/%Y") if row["Date"] is not None else ""
            st.markdown(f'<span style="opacity:{opacity}">{date_str}</span>', unsafe_allow_html=True)
        with col_lib:
            st.markdown(f'<span style="opacity:{opacity}">{libelle_str}</span>', unsafe_allow_html=True)
        with col_debit:
            v = f"**{td:,.2f}**".replace(",", " ") if td else ""
            st.markdown(f'<span style="opacity:{opacity};color:#c0392b">{v}</span>', unsafe_allow_html=True)
        with col_credit:
            v = f"**{tc:,.2f}**".replace(",", " ") if tc else ""
            st.markdown(f'<span style="opacity:{opacity};color:#27ae60">{v}</span>', unsafe_allow_html=True)
        with col_impute:
            # For credit rows show recalculated imputation
            display_impute = impute_str if is_credit and current_active else (str(row.get("Imputé sur", "")) if pd.notna(row.get("Imputé sur")) else "")
            if display_impute in ("nan", "NaN", "None"):
                display_impute = ""
            st.markdown(f'<small style="opacity:{opacity};color:#555">{display_impute}</small>', unsafe_allow_html=True)
        with col_solde:
            if current_active and solde_val is not None:
                color = "#c0392b" if solde_val > 0 else "#27ae60"
                st.markdown(f'<span style="color:{color};font-weight:bold;opacity:{opacity}">{solde_val:,.2f}</span>'.replace(",", " "), unsafe_allow_html=True)
            else:
                st.markdown("")

    # ── Solde final ─────────────────────────────────────────────────
    active_idx_list = [i for i, a in enumerate(active_mask) if a]
    if active_idx_list:
        solde_final = soldes_map.get(max(active_idx_list), 0)
    else:
        solde_final = 0

    st.markdown("---")
    col_s1, col_s2, col_s3 = st.columns([3, 1, 1])
    with col_s2:
        label = "🔴 Solde dû" if solde_final > 0 else "🟢 Crédit"
        st.metric(label, f"{abs(solde_final):,.2f} €".replace(",", " "))
    with col_s3:
        if solde_final > 0:
            st.error(f"Dette : {solde_final:,.2f} €".replace(",", " "))
        else:
            st.success(f"Crédit : {abs(solde_final):,.2f} €".replace(",", " "))

    # ── Boutons d'action rapide ──────────────────────────────────────
    st.markdown("---")
    col_btn1, col_btn2, col_btn3 = st.columns(3)
    with col_btn1:
        if st.button("✅ Tout sélectionner"):
            st.session_state["active_mask"] = [True] * len(df_rel)
            st.rerun()
    with col_btn2:
        if st.button("❌ Tout désélectionner"):
            st.session_state["active_mask"] = [False] * len(df_rel)
            st.rerun()
    with col_btn3:
        if st.button("🔄 Réinitialiser"):
            st.session_state["active_mask"] = [True] * len(df_rel)
            st.session_state["new_rows"] = []
            st.rerun()

    # ── Export Excel ─────────────────────────────────────────────────
    st.markdown("---")
    st.markdown('<div class="section-header">📥 Exporter le relevé recalculé</div>', unsafe_allow_html=True)

    if st.button("🔧 Générer l'export Excel"):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Relevé recalculé"

        # Title
        ws.merge_cells("A1:G1")
        ws["A1"] = f"RELEVÉ DE COMPTE — {copro_name.upper()}"
        ws["A1"].font = Font(bold=True, size=13, color="1a3c6e")
        ws["A1"].alignment = Alignment(horizontal="center")

        ws.merge_cells("A2:G2")
        ws["A2"] = f"Règlement par extinction de la dette la plus ancienne — généré le {datetime.now().strftime('%d/%m/%Y')}"
        ws["A2"].font = Font(italic=True, size=10, color="555555")
        ws["A2"].alignment = Alignment(horizontal="center")

        headers = ["Date", "Libellé", "Débit (€)", "Crédit (€)", "Imputé sur", "Surplus (€)", "Solde (€)"]
        header_fill = PatternFill("solid", start_color="1a3c6e")
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        thin = Side(style="thin", color="CCCCCC")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        row_num = 5
        for i, row in df_rel.iterrows():
            if not active_mask[i]:
                continue
            is_credit = pd.notna(row["Crédit (€)"]) and pd.isna(row["Débit (€)"])
            solde_val = soldes_map.get(i)
            impute_val = imputation_map.get(i, "") if is_credit else ""

            vals = [
                row["Date"].strftime("%d/%m/%Y") if row["Date"] else "",
                row["Libellé"] if not pd.isna(row["Libellé"]) else "",
                row["Débit (€)"] if pd.notna(row["Débit (€)"]) else "",
                row["Crédit (€)"] if pd.notna(row["Crédit (€)"]) else "",
                impute_val,
                "",
                solde_val if solde_val is not None else "",
            ]
            fill_color = "eafaf1" if is_credit else "FFFFFF"
            row_fill = PatternFill("solid", start_color=fill_color)
            for col, val in enumerate(vals, 1):
                cell = ws.cell(row=row_num, column=col, value=val)
                cell.border = border
                cell.fill = row_fill
                if col in (3, 4, 7) and val != "":
                    cell.number_format = "#,##0.00"
                    cell.alignment = Alignment(horizontal="right")
            row_num += 1

        # Solde final row
        ws.cell(row=row_num, column=2, value="SOLDE FINAL").font = Font(bold=True)
        sc = ws.cell(row=row_num, column=7, value=solde_final)
        sc.font = Font(bold=True, color="C0392B" if solde_final > 0 else "27AE60")
        sc.number_format = "#,##0.00"

        # Column widths
        widths = [12, 45, 12, 12, 55, 12, 12]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        st.download_button(
            label="📥 Télécharger le relevé Excel",
            data=buf,
            file_name=f"releve_{copro_name.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )


# ═══════════════════════════════════════════════════════════════════
# TAB 2 — Dettes non soldées
# ═══════════════════════════════════════════════════════════════════
with tab2:
    st.markdown('<div class="section-header">Dettes non soldées (recalculées)</div>', unsafe_allow_html=True)

    # Recalculate remaining debts from active rows
    debts_display = []
    for d in debts_remaining:
        if d["reste"] > 0.005:
            debts_display.append({
                "Date": d["date"].strftime("%d/%m/%Y") if d["date"] else "",
                "Libellé": d["libelle"],
                "Montant initial (€)": d["montant"],
                "Solde restant (€)": round(d["reste"], 2),
            })

    if debts_display:
        df_debts_display = pd.DataFrame(debts_display)
        total_dette = df_debts_display["Solde restant (€)"].sum()

        st.dataframe(
            df_debts_display.style.format({
                "Montant initial (€)": "{:,.2f}",
                "Solde restant (€)": "{:,.2f}",
            }).applymap(
                lambda v: "color: #c0392b; font-weight: bold" if isinstance(v, float) else "",
                subset=["Solde restant (€)"]
            ),
            use_container_width=True,
            hide_index=True,
        )

        st.metric("💰 Total dettes non soldées", f"{total_dette:,.2f} €".replace(",", " "))

        # Export dettes
        if st.button("📥 Exporter les dettes non soldées"):
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter

            wb2 = openpyxl.Workbook()
            ws2 = wb2.active
            ws2.title = "Dettes non soldées"
            ws2.merge_cells("A1:D1")
            ws2["A1"] = "DETTES NON SOLDÉES"
            ws2["A1"].font = Font(bold=True, size=13, color="1a3c6e")
            ws2["A1"].alignment = Alignment(horizontal="center")

            heads = ["Date", "Libellé", "Montant initial (€)", "Solde restant (€)"]
            hfill = PatternFill("solid", start_color="1a3c6e")
            for c, h in enumerate(heads, 1):
                cell = ws2.cell(row=3, column=c, value=h)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = hfill

            for r, d in enumerate(debts_display, 4):
                ws2.cell(row=r, column=1, value=d["Date"])
                ws2.cell(row=r, column=2, value=d["Libellé"])
                ws2.cell(row=r, column=3, value=d["Montant initial (€)"])
                ws2.cell(row=r, column=4, value=d["Solde restant (€)"])

            ws2.column_dimensions["A"].width = 12
            ws2.column_dimensions["B"].width = 50
            ws2.column_dimensions["C"].width = 18
            ws2.column_dimensions["D"].width = 18

            buf2 = io.BytesIO()
            wb2.save(buf2)
            buf2.seek(0)
            st.download_button(
                label="📥 Télécharger",
                data=buf2,
                file_name=f"dettes_{copro_name.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
    else:
        st.success("✅ Aucune dette non soldée ! Compte soldé.")

    # Also show original imported dettes
    if st.session_state["dettes_df"] is not None:
        with st.expander("📄 Dettes originales du fichier importé"):
            df_d_orig = st.session_state["dettes_df"].copy()
            df_d_orig["Date"] = df_d_orig["Date"].apply(lambda x: x.strftime("%d/%m/%Y") if x else "")
            st.dataframe(df_d_orig, use_container_width=True, hide_index=True)


# ═══════════════════════════════════════════════════════════════════
# TAB 3 — Ajouter des lignes
# ═══════════════════════════════════════════════════════════════════
with tab3:
    st.markdown('<div class="section-header">Ajouter une ligne manuellement</div>', unsafe_allow_html=True)
    st.info("Ajoutez des appels ou des règlements supplémentaires. Ils seront intégrés dans le recalcul.")

    with st.form("add_row_form"):
        col_a1, col_a2 = st.columns(2)
        with col_a1:
            new_date = st.date_input("Date", value=date.today())
            new_libelle = st.text_input("Libellé", placeholder="Ex: Appel de fonds 1er trim. 2026")
        with col_a2:
            new_type = st.radio("Type", ["Débit (appel)", "Crédit (règlement)"], horizontal=True)
            new_montant = st.number_input("Montant (€)", min_value=0.0, step=0.01, format="%.2f")

        submitted = st.form_submit_button("➕ Ajouter la ligne")

    if submitted and new_libelle and new_montant > 0:
        new_ts = pd.Timestamp(new_date)
        if new_type == "Débit (appel)":
            new_row = {
                "Date": new_ts,
                "Libellé": new_libelle,
                "Débit (€)": new_montant,
                "Crédit (€)": np.nan,
                "Imputé sur": np.nan,
                "Surplus (€)": np.nan,
                "Solde (€)": np.nan,
            }
        else:
            new_row = {
                "Date": new_ts,
                "Libellé": new_libelle,
                "Débit (€)": np.nan,
                "Crédit (€)": new_montant,
                "Imputé sur": np.nan,
                "Surplus (€)": np.nan,
                "Solde (€)": np.nan,
            }

        # Append to dataframe and re-sort by date
        new_df_row = pd.DataFrame([new_row])
        st.session_state["releve_df"] = pd.concat(
            [st.session_state["releve_df"], new_df_row], ignore_index=True
        ).sort_values("Date", kind="stable").reset_index(drop=True)
        st.session_state["active_mask"] = [True] * len(st.session_state["releve_df"])
        st.success(f"✅ Ligne ajoutée : {new_libelle} — {new_montant:.2f} €")
        st.rerun()

    # Liste des lignes ajoutées récemment
    df_r2 = st.session_state["releve_df"]
    if len(df_r2) > 0:
        st.markdown("---")
        st.markdown("#### Aperçu du relevé complet (trié par date)")
        preview = df_r2[["Date", "Libellé", "Débit (€)", "Crédit (€)"]].copy()
        preview["Date"] = preview["Date"].apply(lambda x: x.strftime("%d/%m/%Y") if x else "")
        st.dataframe(preview, use_container_width=True, hide_index=True)
